import customtkinter as ctk
import math
import random
import openpyxl
import openpyxl.workbook
from tkinter import filedialog
import os
import json
import tkinter.colorchooser as colorchooser

class variable_task_list():
    def __init__(self,name,col_pointer,starting_row):
        self.name = name
        self.col_pointer = col_pointer
        self.starting_row = starting_row
        self.row_pointer = starting_row
        self.button = None
        self.button2 = None
        self.value = []

    def appended(self,value):
        self.value.append(value)

    def add_tasks(self):
        # Generate Events
        self.col_pointer += 1
        if (self.col_pointer > 5):
            # print("Asking to shift down")
            # self.row_pointer += 1
            self.col_pointer = 2
            shift_down(self.row_pointer+1)
        task_textbox = ctk.CTkTextbox(task_frame, width=325, height=100)
        task_textbox.grid(row=self.row_pointer, column=self.col_pointer, padx=5, pady=5)
        task_textbox.insert(0.0,"\n\n\n\nNumber Required: ")
        self.appended(task_textbox)
        self.button.grid(row=self.row_pointer, column=self.col_pointer+1, padx=5, pady=5)
        self.button2.grid(row=self.row_pointer, column=self.col_pointer+1, padx=5, pady=5)
        self.button2.configure(text="Delete a task")
        task_entries[self.starting_row] = self
    
    def remove_tasks(self):
        if len(self.value) > 1:
            self.value.pop().destroy()
            self.col_pointer -= 1
            if (self.col_pointer<2):
                # print("Asking to shift up")
                # self.row_pointer -= 1
                # if self.row_pointer <= 0 : self.row_pointer = 0
                self.col_pointer = 5
                shift_up(self.row_pointer-1)
            self.button.grid(row=self.row_pointer, column=self.col_pointer + 1, padx=5, pady=5)
            self.button2.grid(row=self.row_pointer, column=self.col_pointer + 1, padx=5, pady=5)
            task_entries[self.starting_row] = self
        else: self.button2.configure(text="Must have 1 task")

class taskdata():
    def __init__(self, name):
        self.name = name
        self.pamdata = []
        self.nsmdata = []
        self.ambigdata = []


class task():
    def __init__(self, event,attribution, raw, req):
        self.event = event
        self.attribution = attribution
        self.raw = raw
        self.reqnum = req
        self.original_reqnum = self.reqnum
        self.cleaned = False
        self.clean()
    def clean(self):
        if not self.cleaned:
            raw = self.raw
            rawl = raw.split()
            print(type(raw))
            print(type(rawl))
            if rawl.pop() == "Required:":
                rawl.pop()
                self.cleaned = True
            else:
                rawl.pop()
                rawl.pop()
                self.cleaned = True
            self.raw = " ".join(rawl)
        try:
            self.reqnum = math.ceil(abs(int(self.reqnum)))
        except:
            self.reqnum = len(master_names_list)
        # input(f"{self.reqnum} is a {type(self.reqnum)}")
        
    def reset(self):
        self.reqnum = self.original_reqnum



# File to store user preferences
SETTINGS_FILE = "settings.json"

# Function to load saved settings
def load_settings():
    try:
        with open(SETTINGS_FILE, "r") as file:
            return json.load(file)
    except FileNotFoundError:
        print("File not found")
        return {"bg_color": "#2B2B2B", "fg_color": "#1F6AA5"}  # Default colors

# Function to save settings
def save_settings(settings):
    with open(SETTINGS_FILE, "w") as file:
        json.dump(settings, file, indent=4)

# Function to change background color
def change_bg_color():
    color = colorchooser.askcolor()[1]  # Get HEX color
    if color:
        settings["bg_color"] = color
        save_settings(settings)
        root.configure(bg=color)

# Function to change foreground (button) color
def change_fg_color():
    color = colorchooser.askcolor()[1]
    if color:
        settings["fg_color"] = color
        save_settings(settings)
        for button in button_list:
            button.configure(fg_color=color)


def set_theme():
    ctk.set_appearance_mode("dark")
    if ctk.get_appearance_mode() == "light":
        ctk.set_appearance_mode("dark")
    else:
        ctk.set_appearance_mode("light")

# Moves each of the task rows down
def shift_down(row):
    # print("\n!shifting down")
    
    for item in task_entries:
        # print(f"The name of the row you are trying to move is {item.name}")
        if item.row_pointer >= row-1:
            item.row_pointer += 1
            # print(f"\nThe starting row of {item.name} is {item.row_pointer}\n")
    for widget in task_frame.winfo_children():
        grid_info = widget.grid_info()
        
        if grid_info["row"] >= row:
            
            widget.grid(row=grid_info["row"]+1)

# Moves each of the task rows up
def shift_up(row):
    # print("\n!shifting up")
    
    for item in task_entries:
        if item.row_pointer >= row:
            item.row_pointer -= 1
            if item.row_pointer <= 0 : item.row_pointer = 0
            
        
        # print(f"\nThe starting row of {item.name} is {item.row_pointer}\n")
    for widget in task_frame.winfo_children():
        grid_info = widget.grid_info()
        
        if grid_info["row"] > row:
            
            widget.grid(row=grid_info["row"]-1)

# Function to handle initial event creation
def submit():
    user_input = event_num_entry.get()
    try:
        user_input = int(user_input)
    except ValueError:
        event_num_entry.delete(0, "end")
        label_result.configure(text="Please enter a valid number!", text_color="red")
    else:
        label_result.configure(text="", text_color="black")
        create_event_fields(user_input)
        btn_submit.configure(command=eventUpdate)
        animate_progress(0.1)

# Function to dynamically update events
def eventUpdate():
    user_input = event_num_entry.get()
    try:
        user_input = int(user_input)
    except ValueError:
        event_num_entry.delete(0, "end")
        label_result.configure(text="Please enter a valid number!", text_color="red")
    else:
        difference = user_input - len(event_entries)
        if difference > 0:  # Add more entries
            create_event_fields(difference, start_index=len(event_entries))
        elif difference < 0:  # Remove extra entries
            remove_event_fields(abs(difference))

# Function to create event and time entry fields
def create_event_fields(count, start_index=0):
    for i in range(count):
        row_frame = ctk.CTkFrame(entries_frame)
        row_frame.pack(pady=5, fill="x")

        # Create Event Entry
        event_entry = ctk.CTkEntry(row_frame, placeholder_text=f"Event {start_index + i + 1}")
        event_entry.pack(side="left", padx=5, expand=True, fill="x")

        event_entry.bind("<FocusIn>", lambda event, eid=i: on_focus_in(event, eid))
        event_entry.bind("<FocusOut>", lambda event, eid=i: on_focus_out(event, eid))

        # Create Time Entry
        time_entry = ctk.CTkEntry(row_frame, placeholder_text="Time (ex: 10:00 AM - 11:00 AM)")
        time_entry.pack(side="left", padx=5, expand=True, fill="x")

        # Track the fields and their frames
        event_entries.append(event_entry)
        time_entries.append(time_entry)
        entry_frames.append(row_frame)

# Function to remove event and time fields
def remove_event_fields(count):
    for _ in range(count):
        last_frame = entry_frames.pop()
        event_entries.pop().destroy()
        time_entries.pop().destroy()
        last_frame.destroy()

# Function to display event and time entries
def showentries():
    # for i, (event_entry, time_entry) in enumerate(zip(event_entries, time_entries)):
    #     print(f"Event {i + 1}: {event_entry.get()} at {time_entry.get()}")
    global eventsconfirm 
    eventsconfirm = True
    intialize_tasks()
    animate_progress(0.25)
    switch_tabs()

# Function to handle the names entered in the second tab
def process_names():
    global pam_names_list
    global nsm_names_list
    global master_names_list
    raw_text = pam_names_textbox.get("1.0", "end").strip()
    pam_names_list = [name.strip() for name in raw_text.splitlines() if name.strip()]  # Clean up names
    raw_text = nsm_names_textbox.get("1.0", "end").strip()
    nsm_names_list = [name.strip() for name in raw_text.splitlines() if name.strip()]  # Clean up names
    master_names_list = pam_names_list + nsm_names_list
    label_names_status.configure(text=f"Processed {len(master_names_list)} names!\nGo to the 'Tasks' tab to continue", text_color="green")
    global namesconfirm 
    namesconfirm = True
    animate_progress(0.4)
    switch_tabs()

# Function to create the tasks gui
def intialize_tasks():
    instruction_tab3.destroy()

    # Clear existing grid
    for widget in task_frame.winfo_children():
        widget.destroy()

    # Generate Events
    for row, (event_entry, time_entry) in enumerate(zip(event_entries, time_entries)):
        
        # Create a new list that holds the tasks
        l = variable_task_list(f"{event_entry.get()}\n{time_entry.get()}",2,row)
        # print(f"l.name is {l.name}")
        # print(f"l.value is {l.name}")
        
        # spacing for asthetics
        spacer = ctk.CTkLabel(task_frame, text="",width=100)
        spacer.grid(row=row, column=0, padx=10, pady=5)

        # Title label
        event_label = ctk.CTkLabel(task_frame, text=f"{event_entry.get()}\n{time_entry.get()}", font=("Helvetica", 12, "bold"))
        event_label.grid(row=row, column=1, padx=10, pady=5)
        task_textbox = ctk.CTkTextbox(task_frame, width=325, height=100)
        task_textbox.grid(row=row, column=2, padx=5, pady=5)
        task_textbox.insert(0.0,"\n\n\n\nNumber Required: ")
        l.appended(task_textbox)

        # Add the button for removing tasks
        task_button = ctk.CTkButton(task_frame, text="Delete a task",fg_color=settings["fg_color"])
        task_button.grid(row=row, column=3, padx=5, pady=5, sticky="s,w")
        button_list.append(task_button)

        l.button2 = task_button
        l.button2.configure(command=l.remove_tasks)
        # print("Button configured")


        # Add the button for adding tasks
        task_button = ctk.CTkButton(task_frame, text="Add a task",fg_color=settings["fg_color"])
        task_button.grid(row=row, column=3, padx=5, pady=5, sticky="n,w")
        button_list.append(task_button)

        l.button = task_button
        l.button.configure(command=l.add_tasks)
        # print("Button configured")
        
        task_entries.append(l)
        # print(f"task_entries is now {task_entries}")

# Function to lock in the tasks
def confirmtasks():
    global tasksconfirm 
    tasksconfirm = True
    obtaintasks()
    animate_progress(0.9)
    switch_tabs()

# Function to check everything is done do generate the spreadsheet
def checkconfirm():
    if(namesconfirm and tasksconfirm and eventsconfirm):
        populate_spreadsheet() 
    else:
        instruction_tab4.configure(text="Please process everything before generating.")

# Function to populate the spreadsheet
def populate_spreadsheet():
    if not can_be_generated:
        return
    instruction_tab4.destroy()
    gettasks()
    animate_progress(0.95)
    '''
    master_PAM_tasks = 0
    master_NSM_tasks = 0
    master_all_tasks = 0

    # TODO This is really sus code and idk if it will stay

    for x in task_data_list:
        master_PAM_tasks += len(x.pamdata)
        master_NSM_tasks += len(x.nsmdata)
        master_all_tasks += len(x.ambigdata)
    '''
    # Clear existing grid 
    '''Checks out'''
    for widget in spreadsheet_frame.winfo_children():
        widget.destroy()

    # Generate headers (events)
    for col, (event_entry, time_entry) in enumerate(zip(event_entries, time_entries)):
        header_label = ctk.CTkLabel(spreadsheet_frame, text=f"{event_entry.get()}\n{time_entry.get()}", font=("Helvetica", 12, "bold"))
        header_label.grid(row=0, column=col + 1, padx=5, pady=5)

    # Generate rows (names) and cells
    for row, name in enumerate(pam_names_list):
        # Name label on the left side
        name_label = ctk.CTkLabel(spreadsheet_frame, text=name, font=("Helvetica", 12))
        name_label.grid(row=row + 1, column=0, padx=5, pady=5)

        # Editable cells for each event
        for col in range(len(event_entries)):
            colin = col+1
            cell = ctk.CTkTextbox(spreadsheet_frame, width=200,height=100)
            cell.grid(row=row + 1, column=colin, padx=5, pady=5)
            microlist = task_data_list[col].pamdata + task_data_list[col].ambigdata
            # input(microlist)
            if len(microlist) > 0:
                rand = random.randint(0,len(microlist)-1)
                selection = microlist[rand]
                selection.clean()
                visited_indexes = []
                while selection.reqnum<=0 and len(visited_indexes)<len(microlist):
                    if not rand in visited_indexes:
                        visited_indexes.append(rand)
                    rand = random.randint(0,len(microlist)-1)
                    if not rand in visited_indexes:
                        selection = microlist[rand]
                print(selection.raw)
                if len(visited_indexes)>=len(microlist):
                    cell.insert(0.0,"No task found")
                else:
                    cell.insert(0.0,selection.raw)
                    selection.reqnum -= 1
            else: cell.insert(0.0,"No task found")
    for row, name in enumerate(nsm_names_list):
        # Name label on the left side
        name_label = ctk.CTkLabel(spreadsheet_frame, text=name, font=("Helvetica", 12))
        name_label.grid(row=row + len(pam_names_list) + 1, column=0, padx=5, pady=5)

        # Editable cells for each event
        for col in range(len(event_entries)):
            colin = col+1
            cell = ctk.CTkTextbox(spreadsheet_frame, width=200,height=100)
            cell.grid(row=row + len(pam_names_list) + 1, column=colin, padx=5, pady=5)
            microlist = task_data_list[col].nsmdata + task_data_list[col].ambigdata
            # input(microlist)
            if len(microlist) > 0:
                rand = random.randint(0,len(microlist)-1)
                selection = microlist[rand]
                selection.clean()
                visited_indexes = []
                while selection.reqnum<=0 and len(visited_indexes)<len(microlist):
                    if not rand in visited_indexes:
                        visited_indexes.append(rand)
                    rand = random.randint(0,len(microlist)-1)
                    if not rand in visited_indexes:
                        selection = microlist[rand]
                    # print(visited_indexes)
                print(selection.raw)
                if len(visited_indexes)>=len(microlist):
                    cell.insert(0.0,"No task found")
                else:
                    cell.insert(0.0,selection.raw)
                    selection.reqnum -= 1
            else: cell.insert(0.0,"No task found")


# Function to get the list of stuff.
def obtaintasks():
    for i in task_entries:
        nc = taskdata(i.name)
        for j in i.value:
            rawtext = j.get("1.0", "end").strip()
            if(rawtext.upper().startswith("PAM")):
                words = rawtext.split()
                required = words[-1]
                words.pop(0)
                newtext = " ".join(words)
                # print("this is for PAM")
                t = task(nc,"PAM",newtext,required)
                nc.pamdata.append(t)
            elif(rawtext.upper().startswith("NSM")):
                words = rawtext.split()
                required = words[-1]
                words.pop(0)
                newtext = " ".join(words)
                # print("this is for NSM")
                t = task(nc,"NSM",newtext,required)
                nc.nsmdata.append(t)
            else:
                words = rawtext.split()
                if rawtext != "" and rawtext != "Number Required:":
                    afiche = "all"
                    required = words[-1]
                    if words[0].lower() == "all": 
                        words.pop(0)
                    elif words[0].lower() == "everyone":
                        words.pop(0)
                        afiche = "everyone"
                    newtext = " ".join(words)
                    t = task(nc,afiche,newtext, required)
                    nc.ambigdata.append(t)
        task_data_list.append(nc)
    gettasks()
def gettasks():
    for i in task_data_list:
        
        for j in i.pamdata:
            j.reset()
            print(f"{i.name} has {j.reqnum} for PAM")
        for j in i.nsmdata:
            j.reset()
            print(f"{i.name} has {j.reqnum} for NSM")
        for j in i.ambigdata:
            j.reset()
            print(f"{i.name} has {j.reqnum} for everyone")
    global can_be_generated 
    can_be_generated = True
    


# Function to export as an excel spreadsheet
def export():
    animate_progress(1)
    wb = openpyxl.Workbook()
    ws = wb.active

    # Extract data from the frame
    data = [""]
    for widget in spreadsheet_frame.winfo_children():
        if isinstance(widget, ctk.CTkTextbox) or isinstance(widget, ctk.CTkLabel):  # Modify if needed
            data.append(widget.get("1.0", "end") if hasattr(widget, 'get') else widget.cget("text"))

    # Assume data is arranged in a row-major order (adjust if needed)
    num_columns = len(event_entries)+1  # Set this based on your actual grid
    rows = [data[i:i+num_columns] for i in range(0, len(data)+1, num_columns)]

    # Write to Excel
    for row in rows:
        ws.append(row)

    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx", 
        filetypes=[("Excel files", "*.xlsx")],
        title="Save As"
    )
    
    if file_path:  # If the user didn't cancel
        
        wb.save(file_path)
        os.startfile(file_path)  # Open in Excel (Windows)
        print(f"Excel file saved as '{file_path}'")


# Function to animate progress bar smoothly
def animate_progress(target_value, step=0.001):
    current_value = progress_bar.get()
    if abs(current_value - target_value) > step:
        new_value = current_value + step if current_value < target_value else current_value - step
        progress_bar.set(new_value)
        root.after(4, lambda: animate_progress(target_value))  # Recursively update
    else:
        progress_bar.set(target_value)  # Ensure it reaches exact target


# Function to switch tabs automatically
def switch_tabs():
    current_index = tabs.index(tabview.get())  # Get current tab index
    next_index = (current_index + 1) % 4  # Cycle through tabs (0 → 1 → 2 → 3 → 0)
    tabview.set(tabs[next_index])  # Change the active tab

def event_text_math():
    progress_value = 0.1 + (len(modified_entries) / len(event_entries))/10
    # 10% -> 20%
    animate_progress(progress_value)

def on_focus_in(event, entry_id):
    active_entries.add(entry_id)
    modified_entries.append(event_entries[entry_id])
    # event_text_math()

def on_focus_out(event, entry_id):
    active_entries.discard(entry_id)
    if event_entries[entry_id].get() == '':
        modified_entries.remove(event_entries[entry_id])
    event_text_math()



settings = load_settings()

# Create the main window
root = ctk.CTk()
root.title("NSO MANPOWER SHEET")
root.geometry("900x600")


# Data structures to track widgets and names
event_entries = []
time_entries = []
task_entries = []
task_data_list = []
entry_frames = []
pam_names_list = []
nsm_names_list = []
master_names_list = []
active_entries = set()
modified_entries = []
tabs = ["Events","Names","Tasks","Spreadsheet"]


button_list = []


tasksconfirm = False
eventsconfirm = False
namesconfirm = False
can_be_generated = False

# Add Taskbar
progress_bar = ctk.CTkProgressBar(root)
progress_bar.pack(pady=10, fill="x", padx=20)
progress_bar.set(0)


# Add Tabview
tabview = ctk.CTkTabview(root, width=850, height=550)
tabview.pack(pady=10, padx=10, fill="both", expand=True)

# Tab 1: Event Management
event_tab = tabview.add("Events")

# Title for Tab 1
title = ctk.CTkLabel(event_tab, text="Build a MANPOWER Sheet", text_color="black", font=("Helvetica", 18, "bold"))
title.pack(padx=50, pady=10)

# Instructions for Tab 1
instruction = ctk.CTkLabel(event_tab, text="Input the number of events.", text_color="black", font=("Helvetica", 14, "italic"))
instruction.pack()

# Input field for number of events
event_num_entry = ctk.CTkEntry(event_tab, placeholder_text="Enter a number (ex: 1)", width=200)
event_num_entry.pack()

spacer = ctk.CTkLabel(event_tab, text="")
spacer.pack(pady=20)

# Submit Button
btn_submit = ctk.CTkButton(event_tab, text="Generate Events", command=submit, fg_color=settings["fg_color"])
btn_submit.pack(pady=10)
button_list.append(btn_submit)

# Label to display results or errors
label_result = ctk.CTkLabel(event_tab, text="")
label_result.pack(pady=10)

# Scrollable Frame for events
scrollable_frame = ctk.CTkScrollableFrame(event_tab, width=850, height=400)
scrollable_frame.pack(pady=10, fill="both", expand=True)

entries_frame = ctk.CTkFrame(scrollable_frame)
entries_frame.pack(pady=10, fill="both", expand=True)

# Show Entries Button
btn_show = ctk.CTkButton(scrollable_frame, text="Show Entries", command=showentries, fg_color=settings["fg_color"])
btn_show.pack(side="bottom",pady=10)
button_list.append(btn_show)

# Tab 2: Name Input
names_tab = tabview.add("Names")

# Title for Tab 2
names_title = ctk.CTkLabel(names_tab, text="Enter Names", text_color="black", font=("Helvetica", 18, "bold"))
names_title.pack(pady=10)

# Instructions for Tab 2
instruction = ctk.CTkLabel(names_tab, text="Enter each name on a new line", text_color="black", font=("Helvetica", 14, "italic"))
instruction.pack()

names_scrollable = ctk.CTkScrollableFrame(names_tab, width=850, height=400)
names_scrollable.pack(pady=10, fill="both", expand=True)

pamlbl = ctk.CTkLabel(names_scrollable, text="PAM", text_color="black", font=("Helvetica", 18, "bold"))
pamlbl.pack(pady=(10,0))

# Multiline Textbox for names
pam_names_textbox = ctk.CTkTextbox(names_scrollable, width=500, height=350)
pam_names_textbox.pack(pady=(0,10))

nsmlbl = ctk.CTkLabel(names_scrollable, text="NSM", text_color="black", font=("Helvetica", 18, "bold"))
nsmlbl.pack(pady=(10,0))

# Multiline Textbox for names
nsm_names_textbox = ctk.CTkTextbox(names_scrollable, width=500, height=350)
nsm_names_textbox.pack(pady=(0,10))

# Process Names Button
btn_process_names = ctk.CTkButton(names_scrollable, text="Process Names", command=process_names, fg_color=settings["fg_color"])
btn_process_names.pack(pady=10)
button_list.append(btn_process_names)

# Label to show the status of name processing
label_names_status = ctk.CTkLabel(names_scrollable, text="")
label_names_status.pack(pady=10)

# Tab 3: Task List
task_tab = tabview.add("Tasks")

# Title for Tab 3
task_title = ctk.CTkLabel(task_tab, text="Tasks", text_color="black", font=("Helvetica", 18, "bold"))
task_title.pack(pady=10)

# Instructions for tab 3
instruction_tab3 = ctk.CTkLabel(task_tab, text="Please process events in the previous tab", text_color="black", font=("Helvetica", 14, "italic"))
instruction_tab3.pack()

# Scrollable frame for Tab 3
task_scrollable = ctk.CTkScrollableFrame(task_tab, width=850, height=400)
task_scrollable.pack(pady=10, fill="both", expand=True)
task_frame = ctk.CTkFrame(task_scrollable)
task_frame.pack(pady=10, fill="both", expand=True)

btn_process_tasks = ctk.CTkButton(task_scrollable, text="Process tasks", command=confirmtasks, fg_color=settings["fg_color"])
btn_process_tasks.pack(pady=10)
button_list.append(btn_process_tasks)



# Tab 4: Name Display (Spreadsheet)
spreadsheet_tab = tabview.add("Spreadsheet")

# Title for Tab 4
spreadsheet_title = ctk.CTkLabel(spreadsheet_tab, text="Spreadsheet View", text_color="black", font=("Helvetica", 18, "bold"))
spreadsheet_title.pack(pady=10)

# Instructions for Tab 4
instruction_tab4 = ctk.CTkLabel(spreadsheet_tab, text="Please process events, names, and tasks in the previous tabs", text_color="black", font=("Helvetica", 14, "italic"))
instruction_tab4.pack()

btn_process_generate = ctk.CTkButton(spreadsheet_tab, text="Generate", command=checkconfirm, fg_color=settings["fg_color"])
btn_process_generate.pack(pady=10)
button_list.append(btn_process_generate)

# Scrollable Frame for spreadsheet
spreadsheet_scrollable = ctk.CTkScrollableFrame(spreadsheet_tab, width=850, height=400)
spreadsheet_scrollable.pack(pady=10, fill="both", expand=True)

spreadsheet_frame = ctk.CTkFrame(spreadsheet_scrollable)
spreadsheet_frame.pack(pady=10, fill="both", expand=True)

create_sheet_btn = ctk.CTkButton(spreadsheet_scrollable,text="Export as a spreadsheet",command=export, fg_color=settings["fg_color"])
create_sheet_btn.pack(pady=10)
button_list.append(create_sheet_btn)


# Tab 5
settings_tab = tabview.add("Settings")

# Title for Tab 4
settings_title = ctk.CTkLabel(settings_tab, text="Settings", text_color="black", font=("Helvetica", 18, "bold"))
settings_title.pack(pady=10)

# Instructions for Tab 4
instruction_tab5 = ctk.CTkLabel(settings_tab, text="[instructions for how to help and also settings]", text_color="black", font=("Helvetica", 14, "italic"))
instruction_tab5.pack()

btn_color_toggle = ctk.CTkButton(settings_tab, text="Switch themes", command=change_bg_color,fg_color=settings["fg_color"])
btn_color_toggle.pack(pady=10)
button_list.append(btn_color_toggle)

btn_fg_toggle = ctk.CTkButton(settings_tab, text="Change fg", fg_color=settings["fg_color"], command=change_fg_color)
btn_fg_toggle.pack(pady=10)
button_list.append(btn_fg_toggle)

# Run the application
root.mainloop()